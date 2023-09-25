from openpyxl import load_workbook
wb = load_workbook('/Users/anastasija.semekooutlook.com/Downloads/1.lab.darbs/test1.xlsx')
ws = wb.active
total = 0

for row in range(2, ws.max_row + 1):
    hour_cell = ws['B' + str(row)]
    rate_cell = ws['C' + str(row)]
    
    if isinstance(hour_cell.value, (int, float)) and isinstance(rate_cell.value, (int, float)):
        salary = hour_cell.value * rate_cell.value
        if salary > 3000:
            total += 1

print(total)
wb.close()